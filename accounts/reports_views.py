"""
Reports — advanced analytics + Excel export.
All views are admin-only.
"""
import io
import json
from datetime import timedelta

from django.contrib.auth.decorators import login_required
from django.db.models import Count, Sum, Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.views import View

from accounts.mixins import AdminRequiredMixin


class ReportsView(AdminRequiredMixin, View):
    template_name = 'accounts/reports.html'

    def get(self, request):
        from members.models import Member, MembershipPlan
        from attendance.models import Attendance
        from billing.models import Payment

        today = timezone.localdate()
        month_start = today.replace(day=1)

        # ── Summary KPIs ────────────────────────────────────────────
        total_revenue = Payment.objects.aggregate(t=Sum('amount'))['t'] or 0
        monthly_revenue = Payment.objects.filter(
            date_paid__gte=month_start
        ).aggregate(t=Sum('amount'))['t'] or 0

        # Last 6 months revenue
        monthly_rev_data = []
        for i in range(5, -1, -1):
            ref = today.replace(day=1) - timedelta(days=i * 28)
            ms = ref.replace(day=1)
            if ms.month == 12:
                me = ms.replace(year=ms.year + 1, month=1, day=1) - timedelta(days=1)
            else:
                me = ms.replace(month=ms.month + 1, day=1) - timedelta(days=1)
            rev = Payment.objects.filter(
                date_paid__gte=ms, date_paid__lte=me
            ).aggregate(t=Sum('amount'))['t'] or 0
            monthly_rev_data.append({
                'month': ms.strftime('%b %Y'),
                'revenue': float(rev),
                'count': Payment.objects.filter(date_paid__gte=ms, date_paid__lte=me).count(),
            })

        # ── Membership plan breakdown ────────────────────────────────
        plan_stats = MembershipPlan.objects.annotate(
            member_count=Count('members'),
            active_count=Count('members', filter=Q(members__status='active')),
        ).values('name', 'price', 'duration_days', 'member_count', 'active_count')

        # ── Attendance stats ─────────────────────────────────────────
        total_checkins = Attendance.objects.count()
        this_month_checkins = Attendance.objects.filter(date__gte=month_start).count()

        # Peak day of week
        from django.db.models.functions import ExtractWeekDay
        peak_dow = (
            Attendance.objects
            .annotate(dow=ExtractWeekDay('date'))
            .values('dow')
            .annotate(cnt=Count('id'))
            .order_by('-cnt')
            .first()
        )
        dow_names = {1: 'Sunday', 2: 'Monday', 3: 'Tuesday', 4: 'Wednesday',
                     5: 'Thursday', 6: 'Friday', 7: 'Saturday'}
        peak_day = dow_names.get(peak_dow['dow'], '—') if peak_dow else '—'

        # Face vs manual ratio
        face_count   = Attendance.objects.filter(method='face').count()
        manual_count = Attendance.objects.filter(method='manual').count()

        # ── Member retention ─────────────────────────────────────────
        active_count  = Member.objects.filter(status='active').count()
        total_members = Member.objects.count()
        retention_rate = round((active_count / total_members * 100), 1) if total_members else 0

        # Members with zero attendance this month
        attended_ids = Attendance.objects.filter(
            date__gte=month_start
        ).values_list('member_id', flat=True).distinct()
        inactive_this_month = Member.objects.filter(
            status='active'
        ).exclude(id__in=attended_ids).count()

        ctx = {
            'today': today,
            'total_revenue': total_revenue,
            'monthly_revenue': monthly_revenue,
            'monthly_rev_data': json.dumps(monthly_rev_data),
            'plan_stats': list(plan_stats),
            'total_checkins': total_checkins,
            'this_month_checkins': this_month_checkins,
            'peak_day': peak_day,
            'face_count': face_count,
            'manual_count': manual_count,
            'retention_rate': retention_rate,
            'inactive_this_month': inactive_this_month,
            'total_members': total_members,
            'active_count': active_count,
        }
        return render(request, self.template_name, ctx)


class ExportReportView(AdminRequiredMixin, View):
    """
    Export Excel report.
    Query param: type = members | attendance | payments | summary
    """

    def get(self, request):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        report_type = request.GET.get('type', 'summary')
        wb = openpyxl.Workbook()

        # Style helpers
        HEADER_FILL = PatternFill('solid', fgColor='4F46E5')
        HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
        ALT_FILL    = PatternFill('solid', fgColor='F0F4FF')
        BORDER      = Border(
            bottom=Side(style='thin', color='E2E8F0'),
        )

        def style_header(ws, headers):
            ws.append(headers)
            for col, _ in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 22

        def auto_width(ws):
            for col in ws.columns:
                max_len = max((len(str(c.value or '')) for c in col), default=10)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

        def alt_rows(ws):
            for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
                if i % 2 == 0:
                    for cell in row:
                        cell.fill = ALT_FILL

        today = timezone.localdate()

        if report_type == 'members':
            from members.models import Member
            ws = wb.active
            ws.title = 'Members'
            style_header(ws, ['Full Name', 'Email', 'Phone', 'Plan', 'Join Date', 'Expiry Date', 'Status'])
            for m in Member.objects.select_related('membership_plan').order_by('full_name'):
                ws.append([
                    m.full_name, m.email, m.phone,
                    m.membership_plan.name, str(m.join_date),
                    str(m.expiry_date), m.get_status_display(),
                ])
            alt_rows(ws)
            auto_width(ws)
            filename = f'members_{today}.xlsx'

        elif report_type == 'attendance':
            from attendance.models import Attendance
            ws = wb.active
            ws.title = 'Attendance'
            style_header(ws, ['Member', 'Date', 'Check-in Time', 'Method'])
            for a in Attendance.objects.select_related('member').order_by('-date', '-check_in_time'):
                ws.append([
                    a.member.full_name, str(a.date),
                    a.check_in_time.strftime('%I:%M %p'), a.get_method_display(),
                ])
            alt_rows(ws)
            auto_width(ws)
            filename = f'attendance_{today}.xlsx'

        elif report_type == 'payments':
            from billing.models import Payment
            ws = wb.active
            ws.title = 'Payments'
            style_header(ws, ['Member', 'Amount (Rs.)', 'Date Paid', 'Period Start', 'Period End', 'Method', 'Notes'])
            for p in Payment.objects.select_related('member').order_by('-date_paid'):
                ws.append([
                    p.member.full_name, float(p.amount), str(p.date_paid),
                    str(p.period_start), str(p.period_end),
                    p.get_payment_method_display(), p.notes,
                ])
            alt_rows(ws)
            auto_width(ws)
            filename = f'payments_{today}.xlsx'

        else:  # summary
            from members.models import Member
            from attendance.models import Attendance
            from billing.models import Payment

            # Sheet 1 — Overview
            ws = wb.active
            ws.title = 'Overview'
            ws['A1'] = 'GymApp — Summary Report'
            ws['A1'].font = Font(bold=True, size=14, color='4F46E5')
            ws['A2'] = f'Generated: {today}'
            ws['A2'].font = Font(italic=True, color='64748B')
            ws.append([])

            style_header(ws, ['Metric', 'Value'])
            month_start = today.replace(day=1)
            rows = [
                ('Total Members', Member.objects.count()),
                ('Active Members', Member.objects.filter(status='active').count()),
                ('Expired Members', Member.objects.filter(status='expired').count()),
                ('Suspended Members', Member.objects.filter(status='suspended').count()),
                ('New Members This Month', Member.objects.filter(join_date__gte=month_start).count()),
                ('Total Check-ins (All Time)', Attendance.objects.count()),
                ('Check-ins This Month', Attendance.objects.filter(date__gte=month_start).count()),
                ('Total Revenue (Rs.)', float(Payment.objects.aggregate(t=Sum('amount'))['t'] or 0)),
                ('Revenue This Month (Rs.)', float(Payment.objects.filter(date_paid__gte=month_start).aggregate(t=Sum('amount'))['t'] or 0)),
            ]
            for row in rows:
                ws.append(list(row))
            alt_rows(ws)
            auto_width(ws)

            # Sheet 2 — Monthly Revenue
            ws2 = wb.create_sheet('Monthly Revenue')
            style_header(ws2, ['Month', 'Revenue (Rs.)', 'Payments Count'])
            for i in range(11, -1, -1):
                ref = today.replace(day=1) - timedelta(days=i * 28)
                ms = ref.replace(day=1)
                if ms.month == 12:
                    me = ms.replace(year=ms.year + 1, month=1, day=1) - timedelta(days=1)
                else:
                    me = ms.replace(month=ms.month + 1, day=1) - timedelta(days=1)
                rev = Payment.objects.filter(date_paid__gte=ms, date_paid__lte=me).aggregate(t=Sum('amount'))['t'] or 0
                cnt = Payment.objects.filter(date_paid__gte=ms, date_paid__lte=me).count()
                ws2.append([ms.strftime('%B %Y'), float(rev), cnt])
            alt_rows(ws2)
            auto_width(ws2)

            # Sheet 3 — Members
            ws3 = wb.create_sheet('Members')
            from members.models import Member
            style_header(ws3, ['Full Name', 'Email', 'Phone', 'Plan', 'Join Date', 'Expiry Date', 'Status'])
            for m in Member.objects.select_related('membership_plan').order_by('full_name'):
                ws3.append([m.full_name, m.email, m.phone, m.membership_plan.name,
                             str(m.join_date), str(m.expiry_date), m.get_status_display()])
            alt_rows(ws3)
            auto_width(ws3)

            filename = f'gymapp_report_{today}.xlsx'

        # Stream response
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
